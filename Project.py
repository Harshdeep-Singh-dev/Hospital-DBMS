import re
import pyinputplus as pyip
from openpyxl import load_workbook
import time

#Verify User is used to validate the user name and password
def verifyUser():
    #For documenting the time..
    startTime = time.time()  

    #Usernames and corresponding Passwords
    userList = ['emp1', 'emp2', 'emp3']
    passwordList = ['12345', 'abcdef', 'ab12cd']

    nameOfUser = input("What is your Name: ")

    #Check if name exists in system
    if nameOfUser in userList:
        indexOfUserLoggingIn = userList.index(nameOfUser)
        passwordForLogin = input("Please Enter Your Password: ")

        #Check if the username has same password it should have
        if passwordForLogin == passwordList[indexOfUserLoggingIn]:
            print("Login successful!")
            endTime = time.time()
            print(f"Time taken for login: {endTime - startTime:.4f} seconds")  # Display elapsed time
            return True
        else:
            print("Invalid Password")
            endTime = time.time()
            print(f"Time taken for login attempt: {endTime - startTime:.4f} seconds")
            return False
    else:
        print("Invalid User Name")
        endTime = time.time()
        print(f"Time taken for login attempt: {endTime - startTime:.4f} seconds")
        return False

#This function validates SQL Query using Regex
def validateSqlQuery(query):
    pattern = re.compile(r"^SELECT\s+([a-zA-Z_][a-zA-Z0-9_]*)\s+WHERE\s+([a-zA-Z_][a-zA-Z0-9_]*)\s*=\s*([\w\s\/\-\.:]+)$", re.IGNORECASE)

    if not pattern.match(query):
        raise Exception("Query format is incorrect. Make sure it's in the form 'SELECT <field> WHERE <column> = <value>'.")

#This Function is userd to write the queris and results in the Text document
def logQueryToFile(query, result, filename='query_log.txt'):
    try:
        with open(filename, 'a', encoding='utf-8') as file:
            file.write(f"QUERY: {query}\n")
            file.write("RESULT:\n")
            if result:
                for line in result:
                    file.write(f"- {line}\n")
            else:
                file.write("No matching records found.\n")
            file.write("-" * 40 + "\n")
    except Exception as e:
        print(f"Failed to log query: {e}")

#This function iterates over woorkbook. 
def searchExcel(whatToFind, columnOfCondition, valueOfCondition, filename='Python Project.xlsx'):
    #This is the list that contains the results of queries(if multiple)
    results = []
    #Try Catch block for loading Workbook
    try:
        wb = load_workbook(filename)
        sheet = wb.active

        #To store first entire row
        headers = {str(cell.value).strip(): idx for idx, cell in enumerate(sheet[1])}
        #print(headers)

        #If whatToFind OR column of Condition does not exist
        if whatToFind not in headers or columnOfCondition not in headers:
            print(f"Error: One or both columns not found in the Excel file.")
            return results

        whatToFindCol = headers[whatToFind]
        conditionCol = headers[columnOfCondition]

        #Iterating over table to find matches
        for rowIndex in range(2, sheet.max_row + 1):
            row = sheet[rowIndex - 1]
            cellValue = row[conditionCol].value
            if str(cellValue).strip() == valueOfCondition.strip():
                resultValue = row[whatToFindCol].value
                results.append(str(resultValue))
        # if results found: print results ELSE print no matching records
        if results:
            print(f"{whatToFind} where {columnOfCondition} = {valueOfCondition}:")
            for r in results:
                print(f"- {r}")
        else:
            print("No matching records found.")

        return results
    #Error Handling
    except FileNotFoundError:
        print("Error: Excel file not found.")
        return results
    except Exception as e:
        print(f"An error occurred: {e}")
        return results

#This function accepts and disects the Query
def handleQueries():
    userInput = pyip.inputYesNo("Would you like to perform a query? (yes to continue, no to stop): ")
    #print(userInput)
    if userInput == 'no':
        print("Exiting program.")
        return
    #While loop to repeat everything until user says no.
    while not(userInput == 'no'):
        query = input("QUERY>> ")

        try:
            validateSqlQuery(query)
            print("Valid Query")

            #Record for all the spaces
            indexOfFirstSpace = query.find(' ')
            indexOfSecondSpace = query.find(' ', indexOfFirstSpace + 1)
            indexOfThirdSpace = query.find(' ', indexOfSecondSpace + 1)
            indexOfFourthSpace = query.find(' ', indexOfThirdSpace + 1)
            indexOfFifthSpace = query.find(' ', indexOfFourthSpace + 1)

            #Determing all parameters of the Query
            whatToFind = query[indexOfFirstSpace+1: indexOfSecondSpace]
            columnOfCondition = query[indexOfThirdSpace+1: indexOfFourthSpace]
            valueOfCondition = query[indexOfFifthSpace+1:]

            #print(f"whatToFind: {whatToFind}")
            #print(f"columnOfCondition: {columnOfCondition}")
            #print(f"valueOfCondition: {valueOfCondition}")

            #Calling Functions
            results = searchExcel(whatToFind, columnOfCondition, valueOfCondition)
            logQueryToFile(query, results)

        except Exception as e:
            print(f"Invalid Query: {e}")

        userInput = pyip.inputYesNo("Would you like to perform another query? (yes to continue, no to stop): ")

#Username and Password Validation
if verifyUser():
    print("Database Access Granted...")
    handleQueries()
else:
    print("Access denied. Exiting program.")
